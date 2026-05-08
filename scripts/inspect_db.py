"""Export SQLite database inspection output to an Excel workbook."""

from __future__ import annotations

import sqlite3
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

DB_PATH = Path("data/netflix.db")
OUTPUT_PATH = Path("data/db_inspection.xlsx")

TABLES = [
    "titles",
    "title_countries",
    "title_genres",
    "title_cast",
    "title_directors",
]


def fetch_rows(connection: sqlite3.Connection, query: str) -> list[tuple[Any, ...]]:
    """Run a query and return all rows."""
    return connection.execute(query).fetchall()


def write_table(
    sheet: Worksheet,
    headers: list[str],
    rows: list[tuple[Any, ...]],
) -> None:
    """Write headers and rows to an Excel sheet."""
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for row in rows:
        sheet.append(row)

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def add_overview_sheet(workbook: Workbook, connection: sqlite3.Connection) -> None:
    """Add a high-level database overview sheet."""
    sheet = workbook.active
    sheet.title = "Overview"

    rows = []

    for table in TABLES:
        count = connection.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
        rows.append((table, count))

    write_table(sheet, ["table_name", "row_count"], rows)


def add_schema_sheet(workbook: Workbook, connection: sqlite3.Connection) -> None:
    """Add schema details for each database table."""
    sheet = workbook.create_sheet("Schema")
    rows = []

    for table in TABLES:
        columns = connection.execute(f"PRAGMA table_info({table})").fetchall()

        for column in columns:
            column_id, name, data_type, not_null, default_value, primary_key = column
            rows.append(
                (
                    table,
                    column_id,
                    name,
                    data_type,
                    bool(not_null),
                    default_value,
                    bool(primary_key),
                )
            )

    write_table(
        sheet,
        [
            "table_name",
            "column_id",
            "column_name",
            "data_type",
            "not_null",
            "default_value",
            "primary_key",
        ],
        rows,
    )


def add_sample_titles_sheet(workbook: Workbook, connection: sqlite3.Connection) -> None:
    """Add a small sample of cleaned title records."""
    rows = fetch_rows(
        connection,
        """
        SELECT show_id, title, type, release_year, rating, duration_value,
               duration_unit, date_added, description
        FROM titles
        ORDER BY show_id
        LIMIT 50
        """,
    )

    sheet = workbook.create_sheet("Sample Titles")
    write_table(
        sheet,
        [
            "show_id",
            "title",
            "type",
            "release_year",
            "rating",
            "duration_value",
            "duration_unit",
            "date_added",
            "description",
        ],
        rows,
    )


def add_top_countries_sheet(workbook: Workbook, connection: sqlite3.Connection) -> None:
    """Add top countries by title count."""
    rows = fetch_rows(
        connection,
        """
        SELECT country, COUNT(*) AS title_count
        FROM title_countries
        GROUP BY country
        ORDER BY title_count DESC
        LIMIT 25
        """,
    )

    sheet = workbook.create_sheet("Top Countries")
    write_table(sheet, ["country", "title_count"], rows)


def add_sarkar_example_sheet(
    workbook: Workbook, connection: sqlite3.Connection
) -> None:
    """Add one joined example showing how normalized tables connect."""
    sheet = workbook.create_sheet("Joined Example")

    title_row = connection.execute(
        """
        SELECT show_id, title, type, release_year, rating, duration_value,
               duration_unit, date_added, description
        FROM titles
        WHERE title = 'Sarkar'
        LIMIT 1
        """
    ).fetchone()

    if title_row is None:
        write_table(sheet, ["message"], [("No Sarkar row found.",)])
        return

    show_id = title_row[0]

    sections = [
        (
            "title",
            [
                "show_id",
                "title",
                "type",
                "release_year",
                "rating",
                "duration_value",
                "duration_unit",
                "date_added",
                "description",
            ],
            [title_row],
        ),
        (
            "countries",
            ["country"],
            fetch_rows(
                connection,
                f"""
                SELECT country
                FROM title_countries
                WHERE show_id = '{show_id}'
                ORDER BY position
                """,
            ),
        ),
        (
            "genres",
            ["genre"],
            fetch_rows(
                connection,
                f"""
                SELECT genre
                FROM title_genres
                WHERE show_id = '{show_id}'
                ORDER BY position
                """,
            ),
        ),
        (
            "directors",
            ["director_name"],
            fetch_rows(
                connection,
                f"""
                SELECT director_name
                FROM title_directors
                WHERE show_id = '{show_id}'
                ORDER BY position
                """,
            ),
        ),
        (
            "cast",
            ["actor_name"],
            fetch_rows(
                connection,
                f"""
                SELECT actor_name
                FROM title_cast
                WHERE show_id = '{show_id}'
                ORDER BY position
                LIMIT 20
                """,
            ),
        ),
    ]

    current_row = 1

    for section_name, headers, rows in sections:
        sheet.cell(row=current_row, column=1, value=section_name)
        sheet.cell(row=current_row, column=1).font = Font(bold=True, size=13)
        current_row += 1

        for column_number, header in enumerate(headers, start=1):
            cell = sheet.cell(row=current_row, column=column_number, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")

        current_row += 1

        for row in rows:
            for column_number, value in enumerate(row, start=1):
                sheet.cell(row=current_row, column=column_number, value=value)
            current_row += 1

        current_row += 2

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        column_letter = column_cells[0].column_letter
        sheet.column_dimensions[column_letter].width = min(max_length + 2, 45)


def export_excel() -> None:
    """Create an Excel workbook that shows the cleaned database structure."""
    if not DB_PATH.exists():
        raise FileNotFoundError("Database not found. Run `python ingest.py` first.")

    workbook = Workbook()

    with sqlite3.connect(DB_PATH) as connection:
        add_overview_sheet(workbook, connection)
        add_schema_sheet(workbook, connection)
        add_sample_titles_sheet(workbook, connection)
        add_top_countries_sheet(workbook, connection)
        add_sarkar_example_sheet(workbook, connection)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)

    print(f"Excel inspection file written: {OUTPUT_PATH}")


if __name__ == "__main__":
    export_excel()
